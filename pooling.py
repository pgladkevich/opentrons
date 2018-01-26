import sys
import os

imp_path = 'C:\\Users\\hwd1crop\\AppData\\Local\\Continuum\\anaconda3\\Lib\\site-packages'
sys.path.append(imp_path)

from opentrons import robot, instruments, containers
import openpyxl


wb = openpyxl.load_workbook('../Inputs/pooling_input.xlsx')
dna = wb.active

data = tuple(dna.columns)

wells = [cell.value for cell in data[0][1:]]
names = [cell.value for cell in data[1][1:]]
concentrations = [cell.value for cell in data [2][1:]]
barcodes = [cell.value for cell in data [3][1:]]
plate = [cell.value for cell in data [4][1:]]

#Find the sample with the highest DNA concentration
highest_conc = max(concentrations)

#Adjust all other samples vol addition to match the conncentration of the highest sample
dna_volumes = [highest_conc / conc for conc in concentrations]
dna_loc_vol = list(zip(wells, plate, dna_volumes))
sorted_loc_vol = sorted(dna_loc_vol, key=lambda x: x[2], reverse=True) 

assert sum(dna_volumes) <1500

number_of_reactions = len(names)

# Define the containers on the deck                              
p20rack_1 = containers.load('tiprack-200ul', 'A3')
p20rack_2 = containers.load('tiprack-200ul', 'B3')
p20rack_3 = containers.load('tiprack-200ul', 'A2')
p20rack_4 = containers.load('tiprack-200ul', 'C2')
samples_1 = containers.load('96-PCR-flat', 'A1')
samples_2 = containers.load('96-PCR-flat', 'B1')
samples_3 = containers.load('96-PCR-flat', 'C1')
trash = containers.load('point', 'D1')
output = containers.load('tube-rack-2ml', 'D3')

p20 = instruments.Pipette(
        axis='b',
        min_volume=2,
        max_volume=20,
        trash_container=trash,
        tip_racks=[p20rack_1, p20rack_2, p20rack_3, p20rack_4]
        )

for well, plate, vol in sorted_loc_vol:

    sample_plate = {1: samples_1 , 2: samples_2, 3: samples_3}

    if vol > 20:
        p20.transfer(
            vol, 
            sample_plate[plate][well].bottom(),
            output['A1'].bottom(), 
            blow_out='true', 
            new_tip='always'
            )
    else:
        p20.pick_up_tip()
        p20.transfer(
            vol,
            sample_plate[plate][well].bottom(),
            output['A1'].bottom(), 
            blow_out='true', 
            new_tip='never'
            )
        p20.drop_tip(home_after=False)

#Record the results
results = wb.create_sheet("pooling_output", 0)

results['A1'].value = "Well"
results['B1'].value = "Sample"
results['C1'].value = "[DNA] (ng/ul)"
results['D1'].value = "Bar code"
results['E1'].value = "DNA volume (ul)"
results['F1'].value = "Origin Plate #"

for i, well in enumerate(wells):
    results.cell(column=1, row=i+2).value = well

for i, name in enumerate(names):
    results.cell(column=2, row=i+2).value = name

for i, concentration in enumerate(concentrations):
    results.cell(column=3, row=i+2).value = concentration

for i, barcode in enumerate(barcodes):
    results.cell(column=4, row=i+2).value = barcode

for i, vol in enumerate(dna_volumes):
    results.cell(column=5, row=i+2).value = vol

for i, plate in enumerate(plate):
    results.cell(column=6, row=i+2).value = plate

wb.save('../Outputs/pooling_output.xlsx')

robot.home()